# app/main.py
from __future__ import annotations
from fastapi import FastAPI, Body, Depends, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Dict, Any, List, Union, Optional
from io import BytesIO
from urllib.parse import quote
import datetime
import sqlite3
import logging
import os

from sqlmodel import Session, select, text

from .database import init_db, get_session, engine
from . import models  # يضمن تسجيل الجداول لدى SQLModel

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# ---------------------- إعداد التطبيق ----------------------
app = FastAPI(title="Alsami Workshop Backend", version="4.0")

# CORS
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],            # افتح مؤقتاً
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# لوج بسيط
logger = logging.getLogger("alsami")
logger.setLevel(logging.INFO)

# ---------------------- أدوات مساعدة ----------------------
def to_dict(obj: Any) -> Dict[str, Any]:
    """متوافق مع SQLModel (Pydantic v1/v2)"""
    if hasattr(obj, "model_dump"):
        return obj.model_dump()
    if hasattr(obj, "dict"):
        return obj.dict()
    return dict(obj)

def _safe_date(s: Optional[str]) -> Optional[datetime.date]:
    if not s:
        return None
    try:
        return datetime.date.fromisoformat(s)
    except Exception:
        return None

def ensure_schema(session: Session) -> Dict[str, Any]:
    """
    تأكد من وجود الأعمدة المطلوبة في كل جدول (SQLite يسمح بإضافة أعمدة).
    لا نغيّر أسماء أعمدة قديمة، فقط نضيف الناقص.
    """
    conn: sqlite3.Connection = session.connection().connection  # raw sqlite3
    cur = conn.cursor()

    # شكل الأعمدة المطلوبة لكل جدول
    required = {
        "enginesupply": [
            ("serial", "TEXT", "''"),
            ("engineType", "TEXT", "''"),
            ("model", "TEXT", "''"),
            ("prevSite", "TEXT", "''"),
            ("supDate", "TEXT", "''"),
            ("supplier", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "engineissue": [
            ("serial", "TEXT", "''"),
            ("currSite", "TEXT", "''"),
            ("receiver", "TEXT", "''"),
            ("requester", "TEXT", "''"),
            ("issueDate", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "enginerehab": [
            ("serial", "TEXT", "''"),
            ("rehabber", "TEXT", "''"),
            ("rehabType", "TEXT", "''"),
            ("rehabDate", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "enginecheck": [
            ("serial", "TEXT", "''"),
            ("inspector", "TEXT", "''"),
            # نستخدم description بدل desc (desc كلمة محجوزة)
            ("description", "TEXT", "''"),
            ("checkDate", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "engineupload": [
            ("serial", "TEXT", "''"),
            ("rehabUp", "TEXT", "''"),
            ("checkUp", "TEXT", "''"),
            ("rehabUpDate", "TEXT", "''"),
            ("checkUpDate", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "enginelathe": [
            ("serial", "TEXT", "''"),
            ("lathe", "TEXT", "''"),
            ("latheDate", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "enginepump": [
            ("serial", "TEXT", "''"),
            ("pumpSerial", "TEXT", "''"),
            ("pumpRehab", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "engineelectrical": [
            ("serial", "TEXT", "''"),
            ("etype", "TEXT", "''"),
            ("starter", "TEXT", "''"),
            ("alternator", "TEXT", "''"),
            ("edate", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "generatorsupply": [
            ("code", "TEXT", "''"),
            ("gType", "TEXT", "''"),
            ("model", "TEXT", "''"),
            ("prevSite", "TEXT", "''"),
            ("supDate", "TEXT", "''"),
            ("supplier", "TEXT", "''"),
            ("vendor", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "generatorissue": [
            ("code", "TEXT", "''"),
            ("issueDate", "TEXT", "''"),
            ("receiver", "TEXT", "''"),
            ("requester", "TEXT", "''"),
            ("currSite", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
        "generatorinspect": [
            ("code", "TEXT", "''"),
            ("inspector", "TEXT", "''"),
            ("elecRehab", "TEXT", "''"),
            ("rehabDate", "TEXT", "''"),
            ("rehabUp", "TEXT", "''"),
            ("checkUp", "TEXT", "''"),
            ("notes", "TEXT", "''"),
        ],
    }

    actions: List[str] = []

    # احصل على أسماء الجداول الموجودة
    cur.execute("SELECT name FROM sqlite_master WHERE type='table'")
    tables = {r[0] for r in cur.fetchall()}

    for table, cols in required.items():
        if table not in tables:
            # الجدول غير موجود (سيتم إنشاؤه من SQLModel.init_db)
            actions.append(f"table_missing:{table}")
            continue

        cur.execute(f"PRAGMA table_info({table})")
        have_cols = {r[1] for r in cur.fetchall()}

        for col_name, col_type, default in cols:
            if col_name not in have_cols:
                cur.execute(
                    f"ALTER TABLE {table} ADD COLUMN {col_name} {col_type} DEFAULT {default}"
                )
                actions.append(f"added:{table}.{col_name}")

    conn.commit()
    return {"changed": actions}

# ---------------------- أحداث بدء التشغيل ----------------------
@app.on_event("startup")
def on_startup():
    init_db()  # ينشئ الجداول لو مفقودة
    # حاول إصلاح المخطط
    with Session(engine) as s:
        info = ensure_schema(s)
        if info["changed"]:
            logger.info(f"[schema] fixed: {info['changed']}")

# ---------------------- معالجات أخطاء عامة ----------------------
@app.exception_handler(Exception)
async def general_exception_handler(request: Request, exc: Exception):
    logger.error(f"Unhandled: {request.method} {request.url} -> {exc}")
    return JSONResponse(status_code=500, content={"ok": False, "error": str(exc)})

# ---------------------- صحة الخادم ----------------------
@app.get("/api/health")
def health() -> Dict[str, Any]:
    return {"ok": True, "time": datetime.datetime.utcnow().isoformat()}

# ---------------------- البحث ----------------------
@app.get("/api/search/{key}")
def search_item(key: str, session: Session = Depends(get_session)) -> Dict[str, Any]:
    """
    يرجّع كل السجلات المتعلقة بالرقم (serial للمحركات، code للمولدات)
    """
    result: Dict[str, Any] = {
        "engines": {
            "supply": [],
            "issue": [],
            "rehab": [],
            "check": [],
            "upload": [],
            "lathe": [],
            "pump": [],
            "electrical": [],
        },
        "generators": {
            "supply": [],
            "issue": [],
            "inspect": [],
        },
    }

    # تأكد من المخطط قبل الاستعلام (يحمي من أخطاء الأعمدة المفقودة)
    ensure_schema(session)

    # محركات (حسب serial)
    result["engines"]["supply"] = [to_dict(r) for r in session.exec(
        select(models.EngineSupply).where(models.EngineSupply.serial == key)
    ).all()]

    result["engines"]["issue"] = [to_dict(r) for r in session.exec(
        select(models.EngineIssue).where(models.EngineIssue.serial == key)
    ).all()]

    result["engines"]["rehab"] = [to_dict(r) for r in session.exec(
        select(models.EngineRehab).where(models.EngineRehab.serial == key)
    ).all()]

    result["engines"]["check"] = [to_dict(r) for r in session.exec(
        select(models.EngineCheck).where(models.EngineCheck.serial == key)
    ).all()]

    result["engines"]["upload"] = [to_dict(r) for r in session.exec(
        select(models.EngineUpload).where(models.EngineUpload.serial == key)
    ).all()]

    result["engines"]["lathe"] = [to_dict(r) for r in session.exec(
        select(models.EngineLathe).where(models.EngineLathe.serial == key)
    ).all()]

    result["engines"]["pump"] = [to_dict(r) for r in session.exec(
        select(models.EnginePump).where(models.EnginePump.serial == key)
    ).all()]

    result["engines"]["electrical"] = [to_dict(r) for r in session.exec(
        select(models.EngineElectrical).where(models.EngineElectrical.serial == key)
    ).all()]

    # مولدات (حسب code)
    result["generators"]["supply"] = [to_dict(r) for r in session.exec(
        select(models.GeneratorSupply).where(models.GeneratorSupply.code == key)
    ).all()]

    result["generators"]["issue"] = [to_dict(r) for r in session.exec(
        select(models.GeneratorIssue).where(models.GeneratorIssue.code == key)
    ).all()]

    result["generators"]["inspect"] = [to_dict(r) for r in session.exec(
        select(models.GeneratorInspect).where(models.GeneratorInspect.code == key)
    ).all()]

    return result

# ---------------------- آخر 3 ----------------------
@app.get("/api/last3/engines")
def last3_engines(session: Session = Depends(get_session)) -> Dict[str, Any]:
    ensure_schema(session)
    rows = session.exec(
        select(models.EngineSupply).order_by(models.EngineSupply.id.desc()).limit(3)
    ).all()
    return {"items": [{"serial": r.serial, "prevSite": r.prevSite or ""} for r in rows]}

@app.get("/api/last3/generators")
def last3_generators(session: Session = Depends(get_session)) -> Dict[str, Any]:
    ensure_schema(session)
    rows = session.exec(
        select(models.GeneratorSupply).order_by(models.GeneratorSupply.id.desc()).limit(3)
    ).all()
    return {"items": [{"code": r.code, "prevSite": r.prevSite or ""} for r in rows]}

# ---------------------- مزامنة ----------------------
@app.post("/api/sync/batch")
def sync_batch(
    payload: Dict[str, Any] = Body(...),
    session: Session = Depends(get_session),
) -> Dict[str, Any]:
    """
    body:
    {
      "items": [{"store":"eng_supply","payload":{...}}, ...],
      "returnKey": "111"   # اختياري -> يرجع نتيجة البحث بعد الحفظ
    }
    """
    ensure_schema(session)

    items: List[Dict[str, Any]] = payload.get("items", []) or []
    return_key: Optional[str] = payload.get("returnKey")
    saved = 0

    store_map = {
        "eng_supply": models.EngineSupply,
        "eng_issue": models.EngineIssue,
        "eng_rehab": models.EngineRehab,
        "eng_check": models.EngineCheck,
        "eng_upload": models.EngineUpload,
        "eng_lathe": models.EngineLathe,
        "eng_pump": models.EnginePump,
        "eng_electrical": models.EngineElectrical,
        "gen_supply": models.GeneratorSupply,
        "gen_issue": models.GeneratorIssue,
        "gen_inspect": models.GeneratorInspect,
    }

    for item in items:
        store = (item.get("store") or "").strip()
        data = item.get("payload") or {}
        Model = store_map.get(store)
        if not Model:
            continue

        # خاصّة: معالجة اختلاف الاسم القديم (desc) إلى description
        if Model is models.EngineCheck:
            if "desc" in data and "description" not in data:
                data["description"] = data.pop("desc")

        obj = Model(**data)
        session.add(obj)
        saved += 1

    session.commit()

    if return_key:
        # أعد نتيجة البحث فورًا
        return search_item(return_key, session)

    return {"ok": True, "saved": saved}

# ---------------------- التصدير ----------------------
@app.post("/api/export/xlsx")
def export_xlsx(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    """
    body:
    {
      "scope": "engines"|"generators"|"both",
      "date_from": "YYYY-MM-DD",   # اختياري
      "date_to":   "YYYY-MM-DD",   # اختياري
      "filename": "تقرير.xlsx"     # اختياري
    }
    """
    ensure_schema(session)

    scope = (payload.get("scope") or "both").lower()
    date_from = _safe_date(payload.get("date_from"))
    date_to   = _safe_date(payload.get("date_to"))
    original_filename: str = payload.get("filename") or "report.xlsx"
    safe_filename = "report.xlsx"  # ASCII فقط لتجنب مشكلة اللاتين1
    encoded_name = quote(original_filename)

    def in_range(d: Optional[str]) -> bool:
        if not (date_from or date_to):
            return True
        dd = _safe_date(d)
        if not dd:
            return False
        if date_from and dd < date_from:
            return False
        if date_to and dd > date_to:
            return False
        return True

    wb = Workbook()
    ws_eng = None
    ws_gen = None

    # نمط تصميم
    header_fill = PatternFill("solid", fgColor="2563eb")  # أزرق
    alt_fill = PatternFill("solid", fgColor="f1f5f9")     # رمادي فاتح
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    normal_font = Font(name="Arial")
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    if scope in ("engines", "both"):
        ws_eng = wb.active
        ws_eng.title = "المحركات"
        ws_eng.sheet_view.rightToLeft = True
        eng_headers = ["الرقم التسلسلي","النوع","الموديل","الموقع السابق","تاريخ التوريد","المورّد","ملاحظات"]
        ws_eng.append(eng_headers)
        for i in range(1, len(eng_headers)+1):
            c = ws_eng.cell(row=1, column=i)
            c.fill = header_fill; c.font = header_font; c.alignment = align_right; c.border = thin_border
            ws_eng.column_dimensions[c.column_letter].width = 25

        # فقط من جدول التوريد + داخل المدى
        eng_rows = session.exec(select(models.EngineSupply)).all()
        r = 2
        for e in eng_rows:
            if in_range(e.supDate):
                row = [e.serial, e.engineType, e.model, e.prevSite, e.supDate, e.supplier, e.notes]
                ws_eng.append(row)
                for i in range(1, len(row)+1):
                    c = ws_eng.cell(row=r, column=i)
                    c.font = normal_font; c.alignment = align_right; c.border = thin_border
                    c.fill = alt_fill if r % 2 else white_fill
                r += 1

    if scope in ("generators", "both"):
        if ws_eng is None:
            ws_gen = wb.active
        else:
            ws_gen = wb.create_sheet()
        ws_gen.title = "المولدات"
        ws_gen.sheet_view.rightToLeft = True
        gen_headers = ["الكود","السعة","الموديل","الموقع السابق","تاريخ التوريد","الجهة","المورد","ملاحظات"]
        ws_gen.append(gen_headers)
        for i in range(1, len(gen_headers)+1):
            c = ws_gen.cell(row=1, column=i)
            c.fill = header_fill; c.font = header_font; c.alignment = align_right; c.border = thin_border
            ws_gen.column_dimensions[c.column_letter].width = 25

        gen_rows = session.exec(select(models.GeneratorSupply)).all()
        r = 2
        for g in gen_rows:
            if in_range(g.supDate):
                row = [g.code, g.gType, g.model, g.prevSite, g.supDate, g.supplier, g.vendor, g.notes]
                ws_gen.append(row)
                for i in range(1, len(row)+1):
                    c = ws_gen.cell(row=r, column=i)
                    c.font = normal_font; c.alignment = align_right; c.border = thin_border
                    c.fill = alt_fill if r % 2 else white_fill
                r += 1

    # توقيع
    ws = ws_gen or ws_eng
    if ws:
        last_row = ws.max_row + 2
        sig = ws.cell(row=last_row, column=1)
        sig.value = f"تاريخ التوليد: {datetime.datetime.utcnow().isoformat()}"
        sig.alignment = align_right
        sig.font = Font(italic=True, color="666666", name="Arial")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    headers_resp = {
        # اجعل filename ASCII وتضع العربية في filename* بالنسب المؤية
        "Content-Disposition": f"attachment; filename={safe_filename}; filename*=UTF-8''{encoded_name}"
    }
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )

# ---------------------- إدارة: إصلاح المخطط ----------------------
@app.post("/api/admin/repair")
def admin_repair(session: Session = Depends(get_session)):
    info = ensure_schema(session)
    return {"ok": True, "changed": info["changed"]}

# ---------------------- إدارة: إدخال بيانات تجريبية ----------------------
@app.get("/api/seed")
def seed(session: Session = Depends(get_session)):
    """
    يضيف بيانات بسيطة إن كانت الجداول فاضية.
    """
    ensure_schema(session)

    has = session.exec(select(models.EngineSupply).limit(1)).first()
    if has:
        return {"ok": True, "message": "وجدت بيانات مسبقة، لم أضف جديد."}

    # محركات
    session.add_all([
        models.EngineSupply(serial="111", engineType="Deutz", model="F4L912",
                            prevSite="المخزن", supDate="2025-10-30", supplier="Yemen Mobile", notes="توريد جديد"),
        models.EngineSupply(serial="222", engineType="Perkins", model="404D-22",
                            prevSite="الحديدة", supDate="2025-10-29", supplier="PowerTech", notes="جاهز للاستخدام"),
        models.EngineSupply(serial="333", engineType="Kubota", model="V2203",
                            prevSite="صنعاء", supDate="2025-10-28", supplier="DieselPro", notes="مؤهل"),
    ])
    session.add_all([
        models.EngineIssue(serial="111", currSite="تعز", receiver="المهندس سامي",
                           requester="التشغيل", issueDate="2025-11-01", notes="تم الصرف"),
    ])
    session.add_all([
        models.EngineRehab(serial="333", rehabber="فريق التأهيل", rehabType="إصلاح كامل",
                           rehabDate="2025-11-02", notes="تغيير حلقات وبمب"),
    ])
    session.add_all([
        models.EngineCheck(serial="222", inspector="فريق الفحص", description="فحص حرارة وضغط زيت",
                           checkDate="2025-11-03", notes="ممتاز"),
    ])

    # مولدات
    session.add_all([
        models.GeneratorSupply(code="GEN001", gType="30kVA", model="FG Wilson",
                               prevSite="المستودع", supDate="2025-10-30", supplier="Yemen Mobile",
                               vendor="PowerMax", notes="جديد"),
        models.GeneratorSupply(code="GEN002", gType="20kVA", model="Perkins",
                               prevSite="إب", supDate="2025-10-29", supplier="Yemen Mobile",
                               vendor="EnergyTech", notes="مستخدم"),
        models.GeneratorSupply(code="GEN003", gType="15kVA", model="Deutz",
                               prevSite="تعز", supDate="2025-10-27", supplier="Yemen Mobile",
                               vendor="DieselPro", notes="مؤهل"),
    ])
    session.add_all([
        models.GeneratorIssue(code="GEN001", issueDate="2025-11-01", receiver="موقع ذمار",
                              requester="قسم الطاقة", currSite="ذمار", notes="سليم")
    ])
    session.add_all([
        models.GeneratorInspect(code="GEN002", inspector="فريق الفحص", elecRehab="نعم",
                                rehabDate="2025-11-02", rehabUp="نعم", checkUp="نعم", notes="تم الرفع")
    ])

    session.commit()
    return {"ok": True, "message": "تم إدخال بيانات تجريبية."}

# ---------------------- تشغيل مباشر محلي ----------------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=9000, reload=True)
