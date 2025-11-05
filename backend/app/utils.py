from typing import List, Dict, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import date

def arabic_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.views.sheetView[0].rightToLeft = True

def style_header(row):
    fill = PatternFill("solid", fgColor="E0F2FE")  # سماوي فاتح
    font = Font(bold=True)
    for cell in row:
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")

thin = Side(style="thin", color="999999")
border_all = Border(top=thin, bottom=thin, left=thin, right=thin)

def autofit(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ""
            except:
                val = ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max(12, max_len+2), 45)

def write_rows(ws, headers: List[str], rows: List[List[Optional[str]]]):
    arabic_sheet(ws)
    ws.append(headers)
    style_header(ws[1])
    for r in rows:
        ws.append(r)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=len(headers)):
        for c in row:
            c.alignment = Alignment(vertical="center")
            c.border = border_all
    autofit(ws)

def build_summary_rows(kind: str, data: Dict[str, Optional[str]]) -> List[List[str]]:
    """
    يحول قاموس العناصر إلى صف واحد للاكسل للبحث الموحّد
    """
    headers = [
        "النوع","الرقم","نوع/مودل","الموقع السابق","الموقع الحالي","المؤهل","الفاحص",
        "رفع مؤهل/فحص","آخر قطع","تاريخ التوريد","تاريخ الصرف","تاريخ التأهيل","تاريخ الفحص"
    ]
    # يمكن استخدام هذه الدالة إذا رغبت بتوحيد التنسيق
    key = data.get("key","")
    row = [
        kind,
        key,
        f"{data.get('supply_type','')}/{data.get('supply_model','')}".strip("/"),
        data.get("supply_prev_site",""),
        data.get("issue_current_site",""),
        data.get("rehab_by","") or data.get("rehab_type",""),
        data.get("check_inspector",""),
        f"{data.get('upload_rehab_file','')}/{data.get('upload_check_file','')}".strip("/"),
        data.get("spare_last",""),
        data.get("supply_date",""),
        data.get("issue_date",""),
        data.get("rehab_date",""),
        data.get("check_date",""),
    ]
    return headers, [row]
