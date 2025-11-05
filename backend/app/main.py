# app/main.py
from fastapi import FastAPI, Body, Depends, Request, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Dict, Any, List, Union, Optional
from io import BytesIO
from urllib.parse import quote
import datetime as dt
import os
import sqlite3

from sqlmodel import Session, select
from app.database import init_db, get_session, engine, DB_PATH
from app import models

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

app = FastAPI(title="Alsami Workshop Backend", version="4.0-final")

# ---------------- CORS ----------------
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # افتحها الآن (يمكن تضييقها لاحقًا)
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# ---------------- Utilities ----------------
def _now_iso() -> str:
    return dt.datetime.utcnow().isoformat()

def _pragma_table_info(conn: sqlite3.Connection, table: str) -> List[str]:
    cur = conn.execute(f'PRAGMA table_info("{table}")')
    return [row[1] for row in cur.fetchall()]  # column names

def _ensure_table(conn: sqlite3.Connection, table: str):
    """يضمن وجود الجدول (لو ما وُجد؛ يُنشأ سريعًا بالحد الأدنى)."""
    try:
        conn.execute(f'SELECT 1 FROM "{table}" LIMIT 1')
    except sqlite3.OperationalError:
        # ننشئ نسخة بسيطة من الجدول (مفتاح + حقول أساسية)؛
        # الجداول سيُكمّلها SQLModel.create_all في init_db لاحقًا.
        minimal = {
            "enginesupply":       'CREATE TABLE IF NOT EXISTS enginesupply (id INTEGER PRIMARY KEY, serial TEXT)',
            "engineissue":        'CREATE TABLE IF NOT EXISTS engineissue (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginerehab":        'CREATE TABLE IF NOT EXISTS enginerehab (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginecheck":        'CREATE TABLE IF NOT EXISTS enginecheck (id INTEGER PRIMARY KEY, serial TEXT)',
            "engineupload":       'CREATE TABLE IF NOT EXISTS engineupload (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginelathe":        'CREATE TABLE IF NOT EXISTS enginelathe (id INTEGER PRIMARY KEY, serial TEXT)',
            "enginepump":         'CREATE TABLE IF NOT EXISTS enginepump (id INTEGER PRIMARY KEY, serial TEXT)',
            "engineelectrical":   'CREATE TABLE IF NOT EXISTS engineelectrical (id INTEGER PRIMARY KEY, serial TEXT)',

            "generatorsupply":    'CREATE TABLE IF NOT EXISTS generatorsupply (id INTEGER PRIMARY KEY, code TEXT)',
            "generatorissue":     'CREATE TABLE IF NOT EXISTS generatorissue (id INTEGER PRIMARY KEY, code TEXT)',
            "generatorinspect":   'CREATE TABLE IF NOT EXISTS generatorinspect (id INTEGER PRIMARY KEY, code TEXT)',
        }
        if table in minimal:
            conn.execute(minimal[table])
            conn.commit()

def _ensure_column(conn: sqlite3.Connection, table: str, column: str, coltype: str):
    _ensure_table(conn, table)
    cols = _pragma_table_info(conn, table)
    if column not in cols:
        conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{column}" {coltype}')
        conn.commit()

def _self_heal_schema():
    """
    يصلّح الأعمدة الناقصة بدون تدمير البيانات:
    - enginesupply.engineType
    - enginecheck.desc (بعض النسخ كانت 'description' فقط)
    ...إلخ
    """
    conn = sqlite3.connect(str(DB_PATH))
    conn.execute("PRAGMA foreign_keys=ON")

    try:
        # محركات - توريد
        _ensure_column(conn, "enginesupply", "serial", "TEXT")
        _ensure_column(conn, "enginesupply", "engineType", "TEXT")
        _ensure_column(conn, "enginesupply", "model", "TEXT")
        _ensure_column(conn, "enginesupply", "prevSite", "TEXT")
        _ensure_column(conn, "enginesupply", "supDate", "TEXT")
        _ensure_column(conn, "enginesupply", "supplier", "TEXT")
        _ensure_column(conn, "enginesupply", "notes", "TEXT")

        # محركات - صرف
        _ensure_column(conn, "engineissue", "serial", "TEXT")
        _ensure_column(conn, "engineissue", "currSite", "TEXT")
        _ensure_column(conn, "engineissue", "receiver", "TEXT")
        _ensure_column(conn, "engineissue", "requester", "TEXT")
        _ensure_column(conn, "engineissue", "issueDate", "TEXT")
        _ensure_column(conn, "engineissue", "notes", "TEXT")

        # محركات - تأهيل
        _ensure_column(conn, "enginerehab", "serial", "TEXT")
        _ensure_column(conn, "enginerehab", "rehabber", "TEXT")
        _ensure_column(conn, "enginerehab", "rehabType", "TEXT")
        _ensure_column(conn, "enginerehab", "rehabDate", "TEXT")
        _ensure_column(conn, "enginerehab", "notes", "TEXT")

        # محركات - فحص (desc/description)
        _ensure_column(conn, "enginecheck", "serial", "TEXT")
        _ensure_column(conn, "enginecheck", "inspector", "TEXT")
        _ensure_column(conn, "enginecheck", "checkDate", "TEXT")
        _ensure_column(conn, "enginecheck", "notes", "TEXT")

        cols_check = _pragma_table_info(conn, "enginecheck")
        if "desc" not in cols_check and "description" in cols_check:
            _ensure_column(conn, "enginecheck", "desc", "TEXT")
            conn.execute('UPDATE enginecheck SET "desc" = COALESCE("desc", description)')
            conn.commit()
        else:
            _ensure_column(conn, "enginecheck", "desc", "TEXT")

        # محركات - رفع/مخرطة/بمبات/كهرباء
        _ensure_column(conn, "engineupload", "serial", "TEXT")
        _ensure_column(conn, "engineupload", "rehabUp", "TEXT")
        _ensure_column(conn, "engineupload", "checkUp", "TEXT")
        _ensure_column(conn, "engineupload", "rehabUpDate", "TEXT")
        _ensure_column(conn, "engineupload", "checkUpDate", "TEXT")
        _ensure_column(conn, "engineupload", "notes", "TEXT")

        _ensure_column(conn, "enginelathe", "serial", "TEXT")
        _ensure_column(conn, "enginelathe", "lathe", "TEXT")
        _ensure_column(conn, "enginelathe", "latheDate", "TEXT")
        _ensure_column(conn, "enginelathe", "notes", "TEXT")

        _ensure_column(conn, "enginepump", "serial", "TEXT")
        _ensure_column(conn, "enginepump", "pumpSerial", "TEXT")
        _ensure_column(conn, "enginepump", "pumpRehab", "TEXT")
        _ensure_column(conn, "enginepump", "notes", "TEXT")

        _ensure_column(conn, "engineelectrical", "serial", "TEXT")
        _ensure_column(conn, "engineelectrical", "etype", "TEXT")
        _ensure_column(conn, "engineelectrical", "starter", "TEXT")
        _ensure_column(conn, "engineelectrical", "alternator", "TEXT")
        _ensure_column(conn, "engineelectrical", "edate", "TEXT")
        _ensure_column(conn, "engineelectrical", "notes", "TEXT")

        # مولدات
        _ensure_column(conn, "generatorsupply", "code", "TEXT")
        _ensure_column(conn, "generatorsupply", "gType", "TEXT")
        _ensure_column(conn, "generatorsupply", "model", "TEXT")
        _ensure_column(conn, "generatorsupply", "prevSite", "TEXT")
        _ensure_column(conn, "generatorsupply", "supDate", "TEXT")
        _ensure_column(conn, "generatorsupply", "supplier", "TEXT")
        _ensure_column(conn, "generatorsupply", "vendor", "TEXT")
        _ensure_column(conn, "generatorsupply", "notes", "TEXT")

        _ensure_column(conn, "generatorissue", "code", "TEXT")
        _ensure_column(conn, "generatorissue", "issueDate", "TEXT")
        _ensure_column(conn, "generatorissue", "receiver", "TEXT")
        _ensure_column(conn, "generatorissue", "requester", "TEXT")
        _ensure_column(conn, "generatorissue", "currSite", "TEXT")
        _ensure_column(conn, "generatorissue", "notes", "TEXT")

        _ensure_column(conn, "generatorinspect", "code", "TEXT")
        _ensure_column(conn, "generatorinspect", "inspector", "TEXT")
        _ensure_column(conn, "generatorinspect", "elecRehab", "TEXT")
        _ensure_column(conn, "generatorinspect", "rehabDate", "TEXT")
        _ensure_column(conn, "generatorinspect", "rehabUp", "TEXT")
        _ensure_column(conn, "generatorinspect", "checkUp", "TEXT")
        _ensure_column(conn, "generatorinspect", "notes", "TEXT")

    except Exception as e:
        print("⚠️ schema self-heal error:", e)
    finally:
        conn.close()

def _safe_list(objs):
    return [o.dict(exclude_none=True) for o in objs]

def ensure_schema(session: Session):
    """
    يضمن إنشاء الجداول (SQLModel) ثم يُصلح الأعمدة الناقصة (SQLite).
    """
    try:
        init_db()            # SQLModel.metadata.create_all(engine)
        _self_heal_schema()  # إصلاح أعمدة ناقصة/قديمة
    except Exception as e:
        print("⚠️ ensure_schema error:", e)

# ---------------- Startup ----------------
@app.on_event("startup")
def _on_startup():
    init_db()
    _self_heal_schema()

# ---------------- Health ----------------
@app.get("/api/health")
def health() -> Dict[str, Any]:
    return {"ok": True, "time": _now_iso()}

# ---------------- Seed ----------------
@app.get("/api/seed")
def seed(session: Session = Depends(get_session)):
    """
    يُعبّي قاعدة البيانات ببيانات تجريبية إذا كانت فاضية.
    """
    try:
        ensure_schema(session)
        has_rows = session.exec(select(models.EngineSupply)).first()
        if has_rows:
            return {"ok": True, "msg": "data-exists"}
        # نستدعي seed من الملف الخاص إن وُجد، وإلا نضيف أقل بيانات تجريبية
        try:
            from app.seed_demo_data import seed as do_seed
            do_seed(session)
        except Exception:
            # fallback سريع
            session.add(models.EngineSupply(serial="111", engineType="Deutz", model="F4L912",
                                            prevSite="المخزن", supDate="2025-10-30",
                                            supplier="Yemen Mobile", notes="توريد جديد"))
            session.add(models.GeneratorSupply(code="GEN001", gType="30kVA", model="FG Wilson",
                                               prevSite="المستودع", supDate="2025-10-30",
                                               supplier="Yemen Mobile", vendor="PowerMax", notes="جديد"))
            session.commit()
        return {"ok": True, "msg": "seeded"}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ---------------- Admin Repair/Rebuild ----------------
@app.post("/api/admin/repair")
def admin_repair():
    try:
        _self_heal_schema()
        return {"ok": True}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

@app.post("/api/admin/rebuild")
def admin_rebuild():
    try:
        init_db()
        _self_heal_schema()
        return {"ok": True, "msg": "rebuilt"}
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)}, status_code=500)

# ---------------- Search ----------------
@app.get("/api/search/{key}")
def search_item(key: str, session: Session = Depends(get_session)) -> Dict[str, Any]:
    """
    يعيد نتائج كاملة للرقم التسلسلي للمحركات (serial)
    أو الكود للمولدات (code).
    """
    ensure_schema(session)
    result: Dict[str, Any] = {
        "engines": { "supply": [], "issue": [], "rehab": [], "check": [], "upload": [],
                     "lathe": [], "pump": [], "electrical": [] },
        "generators": { "supply": [], "issue": [], "inspect": [] },
    }
    try:
        # محركات (serial = key)
        result["engines"]["supply"] = _safe_list(
            session.exec(select(models.EngineSupply).where(models.EngineSupply.serial == key)).all()
        )
        result["engines"]["issue"] = _safe_list(
            session.exec(select(models.EngineIssue).where(models.EngineIssue.serial == key)).all()
        )
        result["engines"]["rehab"] = _safe_list(
            session.exec(select(models.EngineRehab).where(models.EngineRehab.serial == key)).all()
        )
        result["engines"]["check"] = _safe_list(
            session.exec(select(models.EngineCheck).where(models.EngineCheck.serial == key)).all()
        )
        result["engines"]["upload"] = _safe_list(
            session.exec(select(models.EngineUpload).where(models.EngineUpload.serial == key)).all()
        )
        result["engines"]["lathe"] = _safe_list(
            session.exec(select(models.EngineLathe).where(models.EngineLathe.serial == key)).all()
        )
        result["engines"]["pump"] = _safe_list(
            session.exec(select(models.EnginePump).where(models.EnginePump.serial == key)).all()
        )
        result["engines"]["electrical"] = _safe_list(
            session.exec(select(models.EngineElectrical).where(models.EngineElectrical.serial == key)).all()
        )

        # مولدات (code = key)
        result["generators"]["supply"] = _safe_list(
            session.exec(select(models.GeneratorSupply).where(models.GeneratorSupply.code == key)).all()
        )
        result["generators"]["issue"] = _safe_list(
            session.exec(select(models.GeneratorIssue).where(models.GeneratorIssue.code == key)).all()
        )
        result["generators"]["inspect"] = _safe_list(
            session.exec(select(models.GeneratorInspect).where(models.GeneratorInspect.code == key)).all()
        )
        return result
    except Exception as e:
        return JSONResponse({"ok": False, "error": f"search-error: {e}"}, status_code=500)

# ---------------- Last 3 ----------------
@app.get("/api/last3/engines")
def last3_engines(session: Session = Depends(get_session)) -> Dict[str, Any]:
    ensure_schema(session)
    try:
        rows = session.exec(
            select(models.EngineSupply).order_by(models.EngineSupply.id.desc()).limit(3)
        ).all()
        return {"items": [{"serial": r.serial, "prevSite": r.prevSite or ""} for r in rows]}
    except Exception as e:
        return {"items": [], "warn": str(e)}

@app.get("/api/last3/generators")
def last3_generators(session: Session = Depends(get_session)) -> Dict[str, Any]:
    ensure_schema(session)
    try:
        rows = session.exec(
            select(models.GeneratorSupply).order_by(models.GeneratorSupply.id.desc()).limit(3)
        ).all()
        return {"items": [{"code": r.code, "prevSite": r.prevSite or ""} for r in rows]}
    except Exception as e:
        return {"items": [], "warn": str(e)}

# ---------------- Sync (IndexedDB) + حفظ مفرد وإرجاع نتيجة مدمجة ----------------
@app.post("/api/sync/batch")
def sync_batch(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    """
    payload = { items: [ {store: 'eng_supply' | 'gen_issue' | ... , payload: {..fields..}} ] }
    يتجاهل أي عنصر خاطئ ويكمل.
    بعد الحفظ يرجّع {"ok":True,"saved":..,"skipped":..}
    """
    ensure_schema(session)
    items: List[Dict[str, Any]] = payload.get("items", [])
    saved, skipped = 0, 0

    STORE_TO_MODEL = {
        # محركات
        "eng_supply": models.EngineSupply,
        "eng_issue": models.EngineIssue,
        "eng_rehab": models.EngineRehab,
        "eng_check": models.EngineCheck,
        "eng_upload": models.EngineUpload,
        "eng_lathe": models.EngineLathe,
        "eng_pump": models.EnginePump,
        "eng_electrical": models.EngineElectrical,
        # مولدات
        "gen_supply": models.GeneratorSupply,
        "gen_issue": models.GeneratorIssue,
        "gen_inspect": models.GeneratorInspect,
    }

    for it in items:
        try:
            store = it.get("store")
            data = it.get("payload", {}) or {}
            Model = STORE_TO_MODEL.get(store)
            if not Model:
                skipped += 1
                continue
            obj = Model(**data)
            session.add(obj)
            saved += 1
        except Exception:
            skipped += 1

    session.commit()
    return {"ok": True, "saved": saved, "skipped": skipped}

@app.post("/api/save-and-search")
def save_and_search(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    """
    لحفظ عنصر واحد ثم إعادة نتيجة بحث مدمجة مباشرة.
    payload = { store: 'eng_supply' | ... , payload: {...}, key: '111' | 'GEN001' }
    """
    ensure_schema(session)
    try:
        STORE_TO_MODEL = {
            "eng_supply": models.EngineSupply,
            "eng_issue": models.EngineIssue,
            "eng_rehab": models.EngineRehab,
            "eng_check": models.EngineCheck,
            "eng_upload": models.EngineUpload,
            "eng_lathe": models.EngineLathe,
            "eng_pump": models.EnginePump,
            "eng_electrical": models.EngineElectrical,
            "gen_supply": models.GeneratorSupply,
            "gen_issue": models.GeneratorIssue,
            "gen_inspect": models.GeneratorInspect,
        }
        store = payload.get("store")
        data = payload.get("payload", {}) or {}
        key = payload.get("key") or data.get("serial") or data.get("code")
        Model = STORE_TO_MODEL.get(store)
        if not Model:
            raise HTTPException(status_code=400, detail="invalid store")
        session.add(Model(**data))
        session.commit()
        # رجّع نتيجة البحث فورًا
        return search_item(str(key), session)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"save-and-search error: {e}")

# ---------------- Export XLSX (قديم + جديد) ----------------
@app.post("/api/export/xlsx")
def export_xlsx(payload: Dict[str, Any] = Body(...), session: Session = Depends(get_session)):
    ensure_schema(session)

    # كشف الوضع: قديم (headers+rows) أم جديد (scope/date range)
    legacy_headers = payload.get("headers")
    legacy_rows    = payload.get("rows")

    original_filename: str = payload.get("filename") or "report.xlsx"
    safe_filename = "report.xlsx"  # ASCII fallback
    encoded_name = quote(original_filename)

    wb = Workbook()
    ws = wb.active
    ws.sheet_view.rightToLeft = True

    # نمط تصميم
    header_fill = PatternFill("solid", fgColor="2563eb")   # أزرق
    alt_fill    = PatternFill("solid", fgColor="f1f5f9")   # صف باهت
    white_fill  = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    normal_font = Font(name="Arial")
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    try:
        if legacy_headers is not None and legacy_rows is not None:
            # ========= الوضع القديم =========
            ws.title = payload.get("sheet") or "تقرير"
            if legacy_headers:
                ws.append(legacy_headers)
                for i in range(1, len(legacy_headers)+1):
                    c = ws.cell(row=1, column=i)
                    c.fill = header_fill; c.font = header_font; c.alignment = align_right; c.border = thin_border
                    ws.column_dimensions[c.column_letter].width = 25
            r = 2
            for row in legacy_rows:
                if isinstance(row, dict) and legacy_headers:
                    ordered = [row.get(h, "") for h in legacy_headers]
                elif isinstance(row, (list, tuple)):
                    ordered = list(row)
                else:
                    ordered = [str(row)]
                ws.append(ordered)
                for i in range(1, len(ordered)+1):
                    c = ws.cell(row=r, column=i)
                    c.font = normal_font; c.alignment = align_right; c.border = thin_border
                    c.fill = alt_fill if r % 2 else white_fill
                r += 1

        else:
            # ========= الوضع الجديد =========
            scope = (payload.get("scope") or "both").lower()
            from_date = payload.get("date_from")
            to_date   = payload.get("date_to")

            def _d(s: Optional[str]):
                try:
                    return dt.date.fromisoformat(s) if s else None
                except Exception:
                    return None

            d1, d2 = _d(from_date), _d(to_date)

            def in_range(iso_str: Optional[str]) -> bool:
                if not (d1 or d2):
                    return True
                try:
                    dd = dt.date.fromisoformat(iso_str) if iso_str else None
                except Exception:
                    return False
                if not dd:
                    return False
                if d1 and dd < d1:
                    return False
                if d2 and dd > d2:
                    return False
                return True

            if scope in ("engines", "both"):
                ws.title = "المحركات"
                eng_headers = ["الرقم التسلسلي","النوع","الموديل","الموقع السابق","تاريخ التوريد","المورّد","ملاحظات"]
                ws.append(eng_headers)
                for i in range(1, len(eng_headers)+1):
                    c = ws.cell(row=1, column=i)
                    c.fill = header_fill; c.font = header_font; c.alignment = align_right; c.border = thin_border
                    ws.column_dimensions[c.column_letter].width = 25
                r = 2
                for e in session.exec(select(models.EngineSupply)).all():
                    if in_range(e.supDate):
                        row = [e.serial, e.engineType, e.model, e.prevSite, e.supDate, e.supplier, e.notes]
                        ws.append(row)
                        for i in range(1, len(row)+1):
                            c = ws.cell(row=r, column=i)
                            c.font = normal_font; c.alignment = align_right; c.border = thin_border
                            c.fill = alt_fill if r % 2 else white_fill
                        r += 1

            if scope in ("generators", "both"):
                sheet = wb.create_sheet("المولدات") if ws.title != "المولدات" else ws
                sheet.sheet_view.rightToLeft = True
                gen_headers = ["الكود","السعة","الموديل","الموقع السابق","تاريخ التوريد","الجهة","المورد","ملاحظات"]
                sheet.append(gen_headers)
                for i in range(1, len(gen_headers)+1):
                    c = sheet.cell(row=1, column=i)
                    c.fill = header_fill; c.font = header_font; c.alignment = align_right; c.border = thin_border
                    sheet.column_dimensions[c.column_letter].width = 25
                r = 2
                for g in session.exec(select(models.GeneratorSupply)).all():
                    if in_range(g.supDate):
                        row = [g.code, g.gType, g.model, g.prevSite, g.supDate, g.supplier, g.vendor, g.notes]
                        sheet.append(row)
                        for i in range(1, len(row)+1):
                            c = sheet.cell(row=r, column=i)
                            c.font = normal_font; c.alignment = align_right; c.border = thin_border
                            c.fill = alt_fill if r % 2 else white_fill
                        r += 1

        # توقيع
        last = ws.max_row + 2
        sig = ws.cell(row=last, column=1)
        sig.value = f"تاريخ التوليد: {dt.datetime.utcnow().isoformat()}"
        sig.alignment = align_right
        sig.font = Font(italic=True, color="666666", name="Arial")

        buf = BytesIO(); wb.save(buf); buf.seek(0)
        headers_resp = {
            "Content-Disposition": f"attachment; filename={safe_filename}; filename*=UTF-8''{encoded_name}",
            "Access-Control-Expose-Headers": "Content-Disposition"
        }
        return StreamingResponse(
            buf,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers=headers_resp,
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export error: {e}")

# ---------------- Root ----------------
@app.get("/")
def root():
    return {"service": "Workshop API", "status": "running", "static": "/static/index.html"}

# ---------------- Favicon (لمنع 404 في Render) ----------------
@app.get("/favicon.ico")
def favicon():
    return JSONResponse(content={}, status_code=204)

# ---------------- Local run ----------------
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=9000, reload=True)
