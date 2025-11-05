from datetime import date, datetime
from typing import Optional
from sqlmodel import select
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session

from fastapi.responses import StreamingResponse
from ..database import get_session
from ..models import (
    EngineSupply, EngineIssue, EngineRehab, EngineCheck, EngineUpload, EngineLathe, EnginePump, EngineElectrical,
    GenSupply, GenIssue, GenInspect
)
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
import io

router = APIRouter(prefix="/export", tags=["export"])

def parse_date(s: str) -> date:
    """تحويل YYYY-MM-DD إلى كائن تاريخ"""
    return date.fromisoformat(s)

def in_range(field, from_: str, to_: str):
    """إنشاء فلتر المدى الزمني"""
    from_d = parse_date(from_)
    to_d = parse_date(to_)
    return (field >= from_d) & (field <= to_d)

    return None

@router.get("/engines")
def export_engines(
    session: Session = Depends(get_session),
    from_: Optional[str] = Query(None, alias="from"),
    to: Optional[str]   = Query(None, alias="to")
):
  
    q = select(EngineSupply)
    rng = in_range(EngineSupply.date, from_, to)
    if rng is not None:
        q = q.where(rng)
    rows = session.exec(q).all()

    headers = ["النوع","المودل","الرقم التسلسلي","الموقع السابق","المورد","تاريخ التوريد","ملاحظات"]
    data = [[r.type, r.model, r.serial, r.prev_site or "", r.supplier or "", str(r.date or ""), r.notes or ""] for r in rows]

    wb = Workbook(); ws = wb.active; ws.title = "المحركات"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"engines_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@router.get("/generators")
def export_generators(session: Session = Depends(get_session), from_: Optional[str] = Query(None, alias="from"), to: Optional[str] = Query(None, alias="to")):
    q = select(GenSupply)
    rng = in_range(GenSupply.date, from_, to)
    if rng is not None:
        q = q.where(rng)
    rows = session.exec(q).all()

    headers = ["النوع","المودل","الترميز","الموقع السابق","المورد","الجهة","تاريخ التوريد","ملاحظات"]
    data = [[r.type, r.model, r.code, r.prev_site or "", r.supplier or "", r.entity or "", str(r.date or ""), r.notes or ""] for r in rows]

    wb = Workbook(); ws = wb.active; ws.title = "المولدات"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"generators_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@router.get("/engines/issue")
def export_engines_issue(session: Session = Depends(get_session), from_: Optional[str] = Query(None, alias="from"), to: Optional[str] = Query(None, alias="to")):
    q = select(EngineIssue)
    rng = in_range(EngineIssue.date, from_, to)
    if rng is not None:
        q = q.where(rng)
    rows = session.exec(q).all()

    headers = ["الرقم التسلسلي","الموقع الحالي","المستلم","الجهة الطالبة","تاريخ الصرف","ملاحظات"]
    data = [[r.serial, r.current_site or "", r.receiver or "", r.requester or "", str(r.date or ""), r.notes or ""] for r in rows]

    wb = Workbook(); ws = wb.active; ws.title = "الصرف محركات"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"eng_issue_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@router.get("/generators/issue")
def export_generators_issue(session: Session = Depends(get_session), from_: Optional[str] = Query(None, alias="from"), to: Optional[str] = Query(None, alias="to")):
    q = select(GenIssue)
    rng = in_range(GenIssue.date, from_, to)
    if rng is not None:
        q = q.where(rng)
    rows = session.exec(q).all()

    headers = ["الترميز","تاريخ الصرف","المستلم","الجهة الطالبة","الموقع الحالي","ملاحظات"]
    data = [[r.code, str(r.date or ""), r.receiver or "", r.requester or "", r.current_site or "", r.notes or ""] for r in rows]

    wb = Workbook(); ws = wb.active; ws.title = "الصرف مولدات"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"gen_issue_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@router.get("/engines/lathe")
def export_engines_lathe(session: Session = Depends(get_session)):
    rows = session.exec(select(EngineLathe)).all()
    headers = ["الرقم التسلسلي","تأهيل المخرطة","تاريخ التوريد للمخرطة","ملاحظات"]
    data = [[r.serial, r.detail or "", str(r.date or ""), r.notes or ""] for r in rows]
    wb = Workbook(); ws = wb.active; ws.title = "المخرطة"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"lathe_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@router.get("/engines/electrical")
def export_engines_electrical(session: Session = Depends(get_session), from_: Optional[str] = Query(None, alias="from"), to: Optional[str] = Query(None, alias="to")):
    q = select(EngineElectrical)
    rng = in_range(EngineElectrical.date, from_, to)
    if rng is not None:
        q = q.where(rng)
    rows = session.exec(q).all()
    headers = ["الرقم التسلسلي","النوع","سلف","دينمو","تاريخ"]
    data = [[r.serial, r.etype or "", r.self_ or "", r.dyn or "", str(r.date or "")] for r in rows]
    wb = Workbook(); ws = wb.active; ws.title = "الصريمي"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"electrical_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})

@router.get("/engines/pump")
def export_engines_pump(session: Session = Depends(get_session)):
    rows = session.exec(select(EnginePump)).all()
    headers = ["رقم المحرك","رقم البمب","تأهيل البمب","ملاحظات"]
    data = [[r.eng_serial, r.pump_serial or "", r.rehab or "", r.notes or ""] for r in rows]
    wb = Workbook(); ws = wb.active; ws.title = "البمبات والنوزلات"
    write_rows(ws, headers, data)
    bio = BytesIO(); wb.save(bio); bio.seek(0)
    fname = f"pump_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f'attachment; filename=\"{fname}\"'})
