# app/main.py
from fastapi import FastAPI, Body, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, JSONResponse
from typing import Dict, Any, List, Union
from io import BytesIO
from urllib.parse import quote
import datetime
import os

from sqlmodel import Session, select

from sqlmodel import SQLModel
from .database import engine
app = FastAPI(title="Alsami Workshop Backend", version="4.0")
# Endpoint إداري مؤقت (احذفه بعد الانتهاء)


from .database import init_db, get_session, DB_PATH
from . import models
try:
    from .seed_demo_data import seed
except ImportError:
    seed = None
  # نزرع بيانات إذا مافيش

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side



# ========== CORS ==========
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],   # افتحها للـ frontend
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ========== STARTUP ==========
@app.on_event("startup")
def on_startup():
    init_db()                # ينشئ الجداول إن لم توجد
    ensure_sqlite_columns()  # يضيف 'desc' إن كان مفقودًا


def ensure_sqlite_columns():
    """
    ترقية خفيفة للسكيو لايت: لو عمود 'desc' ناقص في enginecheck نضيفه.
    """
    with engine.connect() as conn:
        rows = conn.exec_driver_sql("PRAGMA table_info(enginecheck);").fetchall()
        cols = {r[1] for r in rows}  # اسم العمود في الفهرس 1
        if "desc" not in cols:
            conn.exec_driver_sql("ALTER TABLE enginecheck ADD COLUMN desc TEXT;")

@app.post("/api/admin/rebuild")
def admin_rebuild():
    SQLModel.metadata.drop_all(engine)
    SQLModel.metadata.create_all(engine)
    return {"ok": True, "msg": "Rebuilt DB schema."}
# ========== HEALTH ==========
@app.get("/api/health")
def health() -> Dict[str, Any]:
    return {
        "ok": True,
        "time": datetime.datetime.utcnow().isoformat(),
        "db_exists": DB_PATH.exists(),
        "db_path": str(DB_PATH),
    }


# =======================================================
# 1) البحث
# =======================================================
@app.get("/api/search/{key}")
def search_item(key: str, session: Session = Depends(get_session)) -> Dict[str, Any]:
    """
    يرجّع كل البيانات المرتبطة بالرقم التسلسلي (محرك)
    أو بالترميز (مولد)
    """
    try:
        # هيكل فاضي
        result: Dict[str, Any] = {
            "engines": {
                "supply": [],
                "issue": [],
                "rehab": [],
                "check": [],
                "upload": [],
                "lathe": [],
                "pump": [],
                "electrical": [],
            },
            "generators": {
                "supply": [],
                "issue": [],
                "inspect": [],
            },
        }

        # --- محركات (serial) ---
        eng_sup = session.exec(
            select(models.EngineSupply).where(models.EngineSupply.serial == key)
        ).all()
        result["engines"]["supply"] = [r.model_dump() for r in eng_sup]

        eng_issue = session.exec(
            select(models.EngineIssue).where(models.EngineIssue.serial == key)
        ).all()
        result["engines"]["issue"] = [r.model_dump() for r in eng_issue]

        eng_rehab = session.exec(
            select(models.EngineRehab).where(models.EngineRehab.serial == key)
        ).all()
        result["engines"]["rehab"] = [r.model_dump() for r in eng_rehab]

        eng_check = session.exec(
            select(models.EngineCheck).where(models.EngineCheck.serial == key)
        ).all()
        result["engines"]["check"] = [r.model_dump() for r in eng_check]

        eng_upload = session.exec(
            select(models.EngineUpload).where(models.EngineUpload.serial == key)
        ).all()
        result["engines"]["upload"] = [r.model_dump() for r in eng_upload]

        eng_lathe = session.exec(
            select(models.EngineLathe).where(models.EngineLathe.serial == key)
        ).all()
        result["engines"]["lathe"] = [r.model_dump() for r in eng_lathe]

        eng_pump = session.exec(
            select(models.EnginePump).where(models.EnginePump.serial == key)
        ).all()
        result["engines"]["pump"] = [r.model_dump() for r in eng_pump]

        eng_elec = session.exec(
            select(models.EngineElectrical).where(models.EngineElectrical.serial == key)
        ).all()
        result["engines"]["electrical"] = [r.model_dump() for r in eng_elec]

        # --- مولدات (code) ---
        gen_sup = session.exec(
            select(models.GeneratorSupply).where(models.GeneratorSupply.code == key)
        ).all()
        result["generators"]["supply"] = [r.model_dump() for r in gen_sup]

        gen_issue = session.exec(
            select(models.GeneratorIssue).where(models.GeneratorIssue.code == key)
        ).all()
        result["generators"]["issue"] = [r.model_dump() for r in gen_issue]

        gen_inspect = session.exec(
            select(models.GeneratorInspect).where(models.GeneratorInspect.code == key)
        ).all()
        result["generators"]["inspect"] = [r.model_dump() for r in gen_inspect]

        return result
    except Exception as e:
        # هنا اللي كان يعطيك 500 على Render
        print("❌ search error:", e)
        return JSONResponse(
            status_code=500,
            content={"error": "search_failed", "detail": str(e)},
        )


# =======================================================
# 1-b) Endpoints للتشخيص
# =======================================================
@app.get("/api/debug/engines")
def debug_engines(session: Session = Depends(get_session)):
    rows = session.exec(select(models.EngineSupply).order_by(models.EngineSupply.id)).all()
    return {"count": len(rows), "items": [r.model_dump() for r in rows]}

@app.get("/api/debug/generators")
def debug_generators(session: Session = Depends(get_session)):
    rows = session.exec(select(models.GeneratorSupply).order_by(models.GeneratorSupply.id)).all()
    return {"count": len(rows), "items": [r.model_dump() for r in rows]}


# =======================================================
# 2) مزامنة الدُفعات من الفرونت
# =======================================================
@app.post("/api/sync/batch")
async def sync_batch(
    payload: Dict[str, Any] = Body(...),
    session: Session = Depends(get_session),
) -> Dict[str, Any]:
    items: List[Dict[str, Any]] = payload.get("items", [])
    saved = 0

    for item in items:
        store = item.get("store")
        data = item.get("payload", {})

        if store == "eng_supply":
            obj = models.EngineSupply(
                serial=data.get("serial", ""),
                engineType=data.get("engineType"),
                model=data.get("model"),
                prevSite=data.get("prevSite"),
                supDate=data.get("supDate"),
                supplier=data.get("supplier"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_issue":
            obj = models.EngineIssue(
                serial=data.get("serial", ""),
                currSite=data.get("currSite"),
                receiver=data.get("receiver"),
                requester=data.get("requester"),
                issueDate=data.get("issueDate"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_rehab":
            obj = models.EngineRehab(
                serial=data.get("serial", ""),
                rehabber=data.get("rehabber"),
                rehabType=data.get("rehabType"),
                rehabDate=data.get("rehabDate"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_check":
            obj = models.EngineCheck(
                serial=data.get("serial", ""),
                inspector=data.get("inspector"),
                desc=data.get("desc"),
                checkDate=data.get("checkDate"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_upload":
            obj = models.EngineUpload(
                serial=data.get("serial", ""),
                rehabUp=data.get("rehabUp"),
                checkUp=data.get("checkUp"),
                rehabUpDate=data.get("rehabUpDate"),
                checkUpDate=data.get("checkUpDate"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_lathe":
            obj = models.EngineLathe(
                serial=data.get("serial", ""),
                lathe=data.get("lathe"),
                latheDate=data.get("latheDate"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_pump":
            obj = models.EnginePump(
                serial=data.get("serial", ""),
                pumpSerial=data.get("pumpSerial"),
                pumpRehab=data.get("pumpRehab"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "eng_electrical":
            obj = models.EngineElectrical(
                serial=data.get("serial", ""),
                etype=data.get("etype"),
                starter=data.get("starter"),
                alternator=data.get("alternator"),
                edate=data.get("edate"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        # -------- مولدات --------
        elif store == "gen_supply":
            obj = models.GeneratorSupply(
                code=data.get("code", ""),
                gType=data.get("gType"),
                model=data.get("model"),
                prevSite=data.get("prevSite"),
                supDate=data.get("supDate"),
                supplier=data.get("supplier"),
                vendor=data.get("vendor"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "gen_issue":
            obj = models.GeneratorIssue(
                code=data.get("code", ""),
                issueDate=data.get("issueDate"),
                receiver=data.get("receiver"),
                requester=data.get("requester"),
                currSite=data.get("currSite"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

        elif store == "gen_inspect":
            obj = models.GeneratorInspect(
                code=data.get("code", ""),
                inspector=data.get("inspector"),
                elecRehab=data.get("elecRehab"),
                rehabDate=data.get("rehabDate"),
                rehabUp=data.get("rehabUp"),
                checkUp=data.get("checkUp"),
                notes=data.get("notes"),
            )
            session.add(obj); saved += 1

    session.commit()
    return {"ok": True, "saved": saved}


# =======================================================
# 3) آخر 3
# =======================================================
@app.get("/api/last3/engines")
def last3_engines(session: Session = Depends(get_session)) -> Dict[str, Any]:
    rows = session.exec(
        select(models.EngineSupply).order_by(models.EngineSupply.id.desc()).limit(3)
    ).all()
    return {
        "items": [
            {"serial": r.serial, "prevSite": r.prevSite or ""} for r in rows
        ]
    }

@app.get("/api/last3/generators")
def last3_generators(session: Session = Depends(get_session)) -> Dict[str, Any]:
    rows = session.exec(
        select(models.GeneratorSupply).order_by(models.GeneratorSupply.id.desc()).limit(3)
    ).all()
    return {
        "items": [
            {"code": r.code, "prevSite": r.prevSite or ""} for r in rows
        ]
    }


# =======================================================
# 4) التصدير
# =======================================================
@app.post("/api/export/xlsx")
async def export_xlsx(payload: Dict[str, Any] = Body(...)):
    original_filename: str = payload.get("filename") or "report.xlsx"
    safe_filename = "report.xlsx"
    sheet_name: str = payload.get("sheet") or "تقرير"
    headers: List[str] = payload.get("headers") or []
    raw_rows: List[Union[Dict[str, Any], List[Any]]] = payload.get("rows") or []

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.rightToLeft = True

    header_fill = PatternFill("solid", fgColor="10b981")
    alt_fill = PatternFill("solid", fgColor="ecfdf3")
    white_fill = PatternFill("solid", fgColor="FFFFFF")
    header_font = Font(bold=True, color="FFFFFF", name="Arial")
    normal_font = Font(name="Arial")
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="DDDDDD"),
        right=Side(style="thin", color="DDDDDD"),
        top=Side(style="thin", color="DDDDDD"),
        bottom=Side(style="thin", color="DDDDDD"),
    )

    if headers:
        ws.append(headers)
        for col_idx, _ in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_right
            cell.border = thin_border
            ws.column_dimensions[cell.column_letter].width = 28

    current_row = 2
    for row in raw_rows:
        if isinstance(row, dict):
            if headers:
                ordered = [row.get(h, "") for h in headers]
            else:
                ordered = list(row.values())
        elif isinstance(row, (list, tuple)):
            ordered = list(row)
        else:
            ordered = [str(row)]

        ws.append(ordered)
        for col_idx in range(1, len(ordered) + 1):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = normal_font
            cell.alignment = align_right
            cell.border = thin_border
            cell.fill = alt_fill if current_row % 2 else white_fill
        current_row += 1

    ws.append([])
    sig_cell = ws.cell(row=current_row + 1, column=1)
    sig_cell.value = f"تم توليد التقرير في: {datetime.datetime.utcnow().isoformat()}"
    sig_cell.alignment = align_right
    sig_cell.font = Font(italic=True, color="555555", name="Arial")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    encoded_name = quote(original_filename)
    headers_resp = {
        "Content-Disposition": (
            f"attachment; filename={safe_filename}; filename*=UTF-8''{encoded_name}"
        )
    }

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers_resp,
    )


# للتشغيل المحلي
if __name__ == "__main__":
    import uvicorn
    uvicorn.run("app.main:app", host="0.0.0.0", port=9000, reload=True)
